"""
Excel Manager — Registro de Trades v6.2
========================================
Genera y actualiza Registro_de_Trades.xlsx desde la Analytics DB (SQLite)
con columnas de estado, tipo de cierre, método, precios USD, Win/Loss y balance acumulado.

Uso:
  - Standalone:   python3 excel_manager.py
  - Desde bot:    from excel_manager import export_trades_to_excel
"""

import sqlite3
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ════════════════════════════════════════════════════════════════
BASE_DIR = Path(__file__).resolve().parent
ANALYTICS_DB = BASE_DIR / "trade_analytics_v62.db"
EXCEL_FILE = BASE_DIR / "Registro_de_Trades.xlsx"

def load_env():
    env_path = BASE_DIR / ".env"
    if env_path.exists():
        import os
        for line in env_path.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            os.environ[key.strip()] = value.strip().strip('"').strip("'")


# ════════════════════════════════════════════════════════════════
# ESTILOS
# ════════════════════════════════════════════════════════════════
# Headers
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Filas de datos
DATA_FONT = Font(name='Calibri', size=10)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

# Colores para resultado
WIN_FILL = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')   # Verde claro
LOSS_FILL = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')   # Rojo claro
OPEN_FILL = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')   # Amarillo claro
NEUTRAL_FILL = PatternFill(start_color='E2E3E5', end_color='E2E3E5', fill_type='solid')# Gris claro

WIN_FONT = Font(name='Calibri', size=10, bold=True, color='155724')    # Verde oscuro
LOSS_FONT = Font(name='Calibri', size=10, bold=True, color='721C24')   # Rojo oscuro
OPEN_FONT = Font(name='Calibri', size=10, bold=True, color='856404')   # Amarillo oscuro

PNL_WIN_FONT = Font(name='Calibri', size=10, bold=True, color='28A745')   # Verde
PNL_LOSS_FONT = Font(name='Calibri', size=10, bold=True, color='DC3545')  # Rojo

# Bordes
THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)

HEADER_BORDER = Border(
    left=Side(style='thin', color='0D1B2A'),
    right=Side(style='thin', color='0D1B2A'),
    top=Side(style='medium', color='0D1B2A'),
    bottom=Side(style='medium', color='0D1B2A'),
)

# Resumen
SUMMARY_FILL = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
SUMMARY_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
SUMMARY_VALUE_FONT = Font(name='Calibri', size=12, bold=True, color='FFD700')

# ════════════════════════════════════════════════════════════════
# COLUMNAS
# ════════════════════════════════════════════════════════════════
HEADERS = [
    'Fecha Apertura',        # A
    'Fecha Cierre',          # B
    'Criptomoneda',          # C
    'Dirección',             # D
    'Apalancamiento',        # E
    'Precio Entrada (USD)',  # F
    'Precio Salida (USD)',   # G
    'Stop Loss (USD)',       # H
    'Estado',                # I
    'Tipo de Cierre',        # J
    'Método de Cierre',      # K
    'Resultado',             # L
    'PnL (USD)',             # M
    'PnL (%)',               # N
    'Comisión (USD)',        # O
    'Balance Acumulado (USD)',# P
    'Duración',              # Q
    'Notas',                 # R
]

COLUMN_WIDTHS = {
    'A': 18, 'B': 18, 'C': 14, 'D': 16, 'E': 15,
    'F': 20, 'G': 20, 'H': 16, 'I': 12, 'J': 18,
    'K': 16, 'L': 14, 'M': 14, 'N': 12, 'O': 14,
    'P': 22, 'Q': 14, 'R': 30,
}


# ════════════════════════════════════════════════════════════════
# HELPERS
# ════════════════════════════════════════════════════════════════
def format_duration(minutes):
    """Convierte minutos float a formato legible."""
    if minutes is None:
        return '—'
    minutes = abs(minutes)
    if minutes < 60:
        return f'{minutes:.0f}m'
    hours = int(minutes // 60)
    mins = int(minutes % 60)
    if hours < 24:
        return f'{hours}h {mins}m'
    days = hours // 24
    hours = hours % 24
    return f'{days}d {hours}h'


def format_price_usd(price):
    """Formatea precio como dólar con precisión dinámica."""
    if price is None:
        return '—'
    if abs(price) < 0.001:
        return f'${price:.8f}'
    elif abs(price) < 1:
        return f'${price:.6f}'
    elif abs(price) < 100:
        return f'${price:.4f}'
    else:
        return f'${price:.2f}'


def format_pnl(pnl):
    """Formatea PnL con signo +/-."""
    if pnl is None:
        return '—'
    sign = '+' if pnl >= 0 else ''
    if abs(pnl) < 0.01:
        return f'{sign}${pnl:.6f}'
    return f'{sign}${pnl:.4f}'


def determine_close_type(trade):
    """Determina el tipo de cierre del trade basado en los campos de la DB."""
    reason = (trade.get('close_reason') or '').upper()
    
    # Prioridad: ver qué TP/SL/BE fue tocado
    if trade.get('hit_sl') and not trade.get('hit_tp1'):
        return 'SL'
    if trade.get('hit_be') and not trade.get('hit_tp1'):
        return 'BE'
    
    # Basado en close_reason
    if 'TP3' in reason:
        return 'TP3'
    if 'TP2' in reason:
        return 'TP2'
    if 'TP1' in reason:
        return 'TP1'
    if 'SL' in reason or 'STOP' in reason:
        return 'SL'
    if 'BE' in reason or 'BREAK' in reason:
        return 'BE'
    if 'MANUAL' in reason:
        return 'Manual'
    if 'TREND' in reason:
        return 'Trend Change'
    if 'EXIT' in reason:
        return 'Exit Signal'
    if reason:
        return reason.title()
    return '—'


def determine_close_method(trade):
    """Determina si fue cierre manual o automático (bot)."""
    reason = (trade.get('close_reason') or '').upper()
    
    if 'MANUAL' in reason:
        return 'Manual'
    # Si tiene close_reason del bot (TP, SL, TREND_CHANGE, EXIT) → Bot
    if any(k in reason for k in ['TP1', 'TP2', 'TP3', 'SL', 'STOP', 'TREND', 'EXIT', 'BE']):
        return 'Bot'
    # Si fue cerrado por el sistema
    if trade.get('close_reason'):
        return 'Bot'
    return '—'


def determine_direction_display(side):
    """Convierte side a formato display."""
    if not side:
        return '—'
    side = side.upper()
    if side in ('BUY', 'LONG'):
        return 'BUY (LONG)'
    elif side in ('SELL', 'SHORT'):
        return 'SELL (SHORT)'
    return side


def build_notes(trade):
    """Construye notas adicionales."""
    parts = []
    conf = trade.get('confidence')
    if conf is not None:
        parts.append(f'Conf: {conf:.0f}%')
    
    tl = trade.get('trend_line')
    if tl:
        parts.append(f'Trend: {tl}')
    
    tf = trade.get('timeframe')
    if tf:
        parts.append(f'TF: {tf}')
    
    # Mostrar TPs alcanzados
    tps_hit = []
    if trade.get('hit_tp1'):
        tps_hit.append('TP1✓')
    if trade.get('hit_tp2'):
        tps_hit.append('TP2✓')
    if trade.get('hit_tp3'):
        tps_hit.append('TP3✓')
    if trade.get('hit_be'):
        tps_hit.append('BE✓')
    if tps_hit:
        parts.append(' '.join(tps_hit))
    
    tid = trade.get('trade_id')
    if tid:
        parts.append(f'ID: {tid}')
    
    return ' | '.join(parts) if parts else ''


# ════════════════════════════════════════════════════════════════
# EXPORTACIÓN PRINCIPAL
# ════════════════════════════════════════════════════════════════
def get_bitget_data():
    """Fetch real-time Equity and Unrealized PnL from Bitget"""
    import os, ccxt
    load_env()
    try:
        exchange = ccxt.bitget({
            'apiKey': os.environ.get("BITGET_API_KEY"),
            'secret': os.environ.get("BITGET_API_SECRET"),
            'password': os.environ.get("BITGET_PASSWORD") or os.environ.get("BG_PASS"),
            'options': {'defaultType': 'swap'},
            'enableRateLimit': True,
        })
        bal = exchange.fetch_balance()
        usdt_total = float(bal.get('USDT', {}).get('total', 0))
        
        poss = exchange.fetch_positions()
        upnl = 0.0
        for p in poss:
            contracts = abs(float(p.get('contracts', 0) or 0))
            if contracts > 0:
                upnl += float(p.get('unrealizedPnl', 0) or 0)
                
        return usdt_total, upnl
    except Exception as e:
        print(f"Error fetching Bitget data: {e}")
        return None, None

def export_trades_to_excel(db_path=None, excel_path=None):
    """
    Exporta todos los trades de la Analytics DB a Excel.
    Se sincroniza con el balance real de Bitget.
    """
    db_path = Path(db_path) if db_path else ANALYTICS_DB
    excel_path = Path(excel_path) if excel_path else EXCEL_FILE
    
    if not db_path.exists():
        print(f"⚠️ No existe la DB: {db_path}")
        return False
    
    try:
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        
        # Obtener todos los trades ordenados cronológicamente
        c.execute('''
            SELECT * FROM trades 
            ORDER BY 
                CASE WHEN opened_at IS NOT NULL THEN opened_at ELSE '9999' END ASC,
                id ASC
        ''')
        trades = [dict(row) for row in c.fetchall()]
        conn.close()
        
        if not trades:
            print("📭 No hay trades en la DB para exportar.")
            return False
        
        # ── Crear workbook ──
        wb = Workbook()
        ws = wb.active
        ws.title = 'Trades'
        
        # ── Escribir headers ──
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER
        
        # ── Congelar headers ──
        ws.freeze_panes = 'A2'
        
        # ── Anchos de columna ──
        for col_letter, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        
        # ── Escribir datos ──
        total_wins = 0
        total_losses = 0
        total_open = 0
        total_pnl = 0.0
        total_commissions = 0.0
        
        bitget_equity, bitget_upnl = get_bitget_data()
        
        for trade in trades:
            if trade.get('status') == 'closed':
                total_pnl += (trade.get('pnl_usdt') or 0.0)
                total_commissions += (trade.get('commission') or 0.0)
        
        if bitget_equity is not None:
            initial_balance = bitget_equity - (total_pnl + total_commissions)
        else:
            initial_balance = 125.97
            
        running_balance = initial_balance
        
        total_wins = 0
        total_losses = 0
        total_be = 0
        total_open = 0
        total_pnl = 0.0
        total_commissions = 0.0
        
        for row_idx, trade in enumerate(trades, 2):
            is_open = trade.get('status') == 'open'
            is_closed = trade.get('status') == 'closed'
            pnl = trade.get('pnl_usdt')
            commission = trade.get('commission') or 0.0
            
            # Excluir artifacts de sync sin PnL (ruido de auditoría)
            if trade.get('close_reason') == 'SYNC_MISSING' and (pnl is None or pnl == 0):
                continue

            is_win = pnl is not None and pnl > 0.0001
            is_loss = pnl is not None and pnl < -0.0001
            is_breakeven = pnl is not None and -0.0001 <= pnl <= 0.0001
            
            # Calcular balance acumulado
            if pnl is not None and is_closed:
                # El balance real toma en cuenta el PnL MÁS las comisiones (que son negativas)
                net_profit = pnl + commission
                running_balance += net_profit
                total_pnl += pnl
                total_commissions += commission
                if is_win:
                    total_wins += 1
                elif is_loss:
                    total_losses += 1
                elif is_breakeven:
                    total_be += 1
            elif is_open:
                total_open += 1
            
            # Determinar campos derivados
            close_type = determine_close_type(trade) if is_closed else '—'
            close_method = determine_close_method(trade) if is_closed else '—'
            direction = determine_direction_display(trade.get('side'))
            notes = build_notes(trade)
            
            # Fecha apertura
            opened_at = trade.get('opened_at', '')
            if opened_at:
                try:
                    dt = datetime.fromisoformat(opened_at)
                    opened_at = dt.strftime('%Y-%m-%d %H:%M')
                except:
                    pass
            
            # Fecha cierre
            closed_at = trade.get('closed_at', '')
            if closed_at:
                try:
                    dt = datetime.fromisoformat(closed_at)
                    closed_at = dt.strftime('%Y-%m-%d %H:%M')
                except:
                    pass
            
            # Resultado display
            if is_open:
                resultado = '⏳ ABIERTA'
                estado = '🟡 ABIERTA'
            elif is_win:
                resultado = '✅ WIN'
                estado = '🔴 CERRADA'
            elif is_loss:
                resultado = '❌ LOSS'
                estado = '🔴 CERRADA'
            else:
                resultado = '➖ B-EVEN'
                estado = '🔴 CERRADA'
            
            # ── Valores de las celdas ──
            row_data = [
                opened_at,                                          # A: Fecha Apertura
                closed_at if is_closed else '—',                    # B: Fecha Cierre
                _format_symbol(trade.get('symbol', '')),            # C: Criptomoneda
                direction,                                          # D: Dirección
                f"{trade.get('leverage', 10)}x",                    # E: Apalancamiento
                format_price_usd(trade.get('entry_price')),         # F: Precio Entrada
                format_price_usd(trade.get('exit_price')) if is_closed else '—',  # G: Precio Salida
                format_price_usd(trade.get('sl_original')),         # H: Stop Loss
                estado,                                             # I: Estado
                close_type,                                         # J: Tipo de Cierre
                close_method,                                       # K: Método
                resultado,                                          # L: Resultado
                format_pnl(pnl) if pnl is not None else '—',       # M: PnL USD
                f"{trade.get('pnl_pct', 0):.2f}%" if trade.get('pnl_pct') is not None else '—',  # N: PnL %
                f"-${abs(trade.get('commission', 0)):.4f}" if trade.get('commission') else '—',  # O: Comisión
                f"${running_balance:.2f}",                          # P: Balance Acumulado
                format_duration(trade.get('duration_min')),         # Q: Duración
                notes,                                              # R: Notas
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                
                # Alineación según columna
                if col_idx in (1, 2, 18):  # Fechas y notas
                    cell.alignment = LEFT_ALIGN
                elif col_idx in (6, 7, 8, 13, 14, 15, 16):  # Precios y PnL
                    cell.alignment = RIGHT_ALIGN
                else:
                    cell.alignment = CENTER_ALIGN
            
            # ── Colorear fila según resultado ──
            if is_open:
                row_fill = OPEN_FILL
                for col_idx in range(1, len(HEADERS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = row_fill
                ws.cell(row=row_idx, column=9).font = OPEN_FONT   # Estado
                ws.cell(row=row_idx, column=12).font = OPEN_FONT  # Resultado
            elif is_win:
                row_fill = WIN_FILL
                for col_idx in range(1, len(HEADERS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = row_fill
                ws.cell(row=row_idx, column=12).font = WIN_FONT   # Resultado
                ws.cell(row=row_idx, column=13).font = PNL_WIN_FONT  # PnL
                ws.cell(row=row_idx, column=9).font = Font(name='Calibri', size=10, color='155724')
            elif is_loss:
                row_fill = LOSS_FILL
                for col_idx in range(1, len(HEADERS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = row_fill
                ws.cell(row=row_idx, column=12).font = LOSS_FONT  # Resultado
                ws.cell(row=row_idx, column=13).font = PNL_LOSS_FONT  # PnL
                ws.cell(row=row_idx, column=9).font = Font(name='Calibri', size=10, color='721C24')
        
        # ── SECCIÓN DE RESUMEN ──
        summary_start = len(trades) + 3  # Dejar una fila vacía
        
        # Línea separadora
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=summary_start, column=col_idx)
            cell.fill = SUMMARY_FILL
            cell.border = HEADER_BORDER
        
        summary_start += 1
        
        summary_data = [
            ('📊 RESUMEN DE PORTAFOLIO', ''),
            ('', ''),
            ('Balance Inicial (Calculado)', f'${initial_balance:.2f}'),
            ('Balance Actual Histórico (Closed)', f'${running_balance:.2f}'),
            ('PnL Abiertos (Unrealized)', f'{("+" if bitget_upnl and bitget_upnl >= 0 else "")}${bitget_upnl:.4f}' if bitget_upnl is not None else 'N/A'),
            ('Equidad Real Bitget (Actual)', f'${bitget_equity:.2f}' if bitget_equity is not None else 'N/A'),
            ('', ''),
            ('PnL Bruto Total', f'{("+" if total_pnl >= 0 else "")}${total_pnl:.4f}'),
            ('Comisiones Totales', f'${total_commissions:.4f}'),
            ('PnL Neto Total', f'{("+" if (total_pnl + total_commissions) >= 0 else "")}${total_pnl + total_commissions:.4f}'),
            ('', ''),
            ('Total Trades Cerrados', f'{total_wins + total_losses + total_be}'),
            ('Trades Ganados', f'{total_wins} ✅'),
            ('Trades Perdidos', f'{total_losses} ❌'),
            ('Trades Empate (BE)', f'{total_be} ➖'),
            ('Trades Abiertos', f'{total_open} ⏳'),
            ('Win Rate (Excl. BE)', f'{(total_wins / (total_wins + total_losses) * 100):.1f}%' if (total_wins + total_losses) > 0 else 'N/A'),
        ]
        
        for i, (label, value) in enumerate(summary_data):
            row = summary_start + i
            label_cell = ws.cell(row=row, column=1, value=label)
            value_cell = ws.cell(row=row, column=2, value=value)
            
            if i == 0:  # Título
                label_cell.font = Font(name='Calibri', size=13, bold=True, color='1B2A4A')
                for col_idx in range(1, 5):
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            else:
                label_cell.font = Font(name='Calibri', size=10, bold=True)
                value_cell.font = SUMMARY_VALUE_FONT if 'Balance' in label or 'PnL' in label else Font(name='Calibri', size=11, bold=True)
                
                # Colorear PnL total
                if 'PnL Total' in label:
                    value_cell.font = Font(name='Calibri', size=12, bold=True, 
                                          color='28A745' if total_pnl >= 0 else 'DC3545')
        
        # ── Guardar ──
        wb.save(str(excel_path))
        print(f"✅ Excel exportado: {excel_path}")
        print(f"   📊 {len(trades)} trades | {total_wins}W / {total_losses}L / {total_open} abiertos")
        print(f"   💰 Balance: ${running_balance:.2f} | PnL: {'+' if total_pnl >= 0 else ''}${total_pnl:.4f}")
        return True
        
    except Exception as e:
        print(f"❌ Error exportando Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


def _format_symbol(symbol):
    """Formatea símbolo para display (BERUSDT → BERA/USDT)."""
    if not symbol:
        return '—'
    s = symbol.upper().replace('.P', '').replace('PERP', '')
    if '/' in s:
        return s
    if s.endswith('USDT'):
        base = s[:-4]
        return f'{base}/USDT'
    return s


# ════════════════════════════════════════════════════════════════
# MIGRACIÓN: Importar datos existentes del Excel viejo a la DB
# ════════════════════════════════════════════════════════════════
def migrate_existing_excel_to_db(excel_path=None, db_path=None):
    """
    Lee el Excel viejo y crea entradas en la Analytics DB para los trades existentes.
    Esto preserva los datos históricos que ya están en el Excel.
    """
    excel_path = Path(excel_path) if excel_path else EXCEL_FILE
    db_path = Path(db_path) if db_path else ANALYTICS_DB
    
    if not excel_path.exists():
        print(f"⚠️ Excel no encontrado: {excel_path}")
        return False
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(excel_path))
        ws = wb['Trades']
        
        # Inicializar DB si no existe
        _ensure_db_tables(db_path)
        
        conn = sqlite3.connect(str(db_path))
        c = conn.cursor()
        
        # Verificar si ya hay datos migrados
        c.execute("SELECT COUNT(*) FROM trades")
        existing_count = c.fetchone()[0]
        if existing_count > 0:
            print(f"ℹ️ La DB ya tiene {existing_count} trades. Saltando migración.")
            conn.close()
            return True
        
        # Agrupar las filas por símbolo para determinar trades completos
        # El Excel viejo tiene operaciones individuales (BUY/SELL por separado)
        # Necesitamos agrupar aperturas y cierres
        trades_dict = {}  # symbol -> {opens: [], closes: []}
        
        for row_idx in range(2, ws.max_row + 1):
            fecha = ws.cell(row=row_idx, column=1).value
            cripto = ws.cell(row=row_idx, column=2).value
            direccion = ws.cell(row=row_idx, column=3).value
            leverage = ws.cell(row=row_idx, column=4).value
            contratos = ws.cell(row=row_idx, column=5).value
            precio = ws.cell(row=row_idx, column=6).value
            resultado = ws.cell(row=row_idx, column=7).value
            comision = ws.cell(row=row_idx, column=8).value
            pnl = ws.cell(row=row_idx, column=9).value
            balance = ws.cell(row=row_idx, column=10).value
            notas = ws.cell(row=row_idx, column=11).value
            
            # Saltar filas vacías, separadores y notas
            if not fecha or not cripto or not direccion:
                continue
            if isinstance(fecha, str) and ('═══' in fecha or 'NOTA:' in fecha or 'Al reiniciar' in fecha):
                continue
            
            is_open_trade = resultado and 'ABIERTA' in str(resultado).upper()
            
            # Para operaciones abiertas, crear trade directamente
            if is_open_trade:
                # Extraer SL de las notas
                sl_val = None
                trade_id = ''
                if notas:
                    import re
                    sl_match = re.search(r'SL:\s*([\d.]+)', str(notas))
                    if sl_match:
                        sl_val = float(sl_match.group(1))
                    tid_match = re.search(r'Trade\s*#(\S+)', str(notas))
                    if tid_match:
                        trade_id = tid_match.group(1)
                
                c.execute('''INSERT INTO trades 
                    (trade_id, symbol, side, entry_price, sl_original, sl_final,
                     leverage, amount_usdt, contracts, opened_at, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'open')''', (
                    trade_id,
                    cripto.replace('/', '').upper() if cripto else '',
                    str(direccion).lower() if direccion else '',
                    float(precio) if precio else None,
                    sl_val,
                    sl_val,
                    int(str(leverage).replace('x', '')) if leverage else 10,
                    10.0,
                    float(contratos) if contratos else None,
                    str(fecha),
                ))
                continue
            
            # Para operaciones históricas, agrupar por cripto+fecha
            key = f"{cripto}_{str(fecha)[:16]}"
            if key not in trades_dict:
                trades_dict[key] = {
                    'symbol': cripto, 'fecha': fecha, 'leverage': leverage,
                    'operations': []
                }
            trades_dict[key]['operations'].append({
                'direccion': direccion, 'contratos': contratos,
                'precio': precio, 'comision': comision, 'pnl': pnl,
            })
        
        # Procesar trades agrupados (históricos)
        for key, group in trades_dict.items():
            ops = group['operations']
            symbol = str(group['symbol']).replace('/', '').upper()
            fecha = str(group['fecha'])
            leverage = group['leverage']
            
            # Calcular PnL total del grupo
            total_pnl = sum(float(op['pnl'] or 0) for op in ops)
            total_comision = sum(float(op['comision'] or 0) for op in ops)
            
            # Determinar side (la primera operación es la apertura)
            first_op = ops[0]
            side = str(first_op['direccion']).lower()
            entry_price = float(first_op['precio']) if first_op['precio'] else None
            
            # Buscar precio de salida (operación opuesta)
            exit_price = None
            for op in ops:
                if str(op['direccion']).upper() != str(first_op['direccion']).upper():
                    exit_price = float(op['precio']) if op['precio'] else None
                    break
            
            # Si no hay operación opuesta, el exit es el último de la misma dir
            if exit_price is None and len(ops) > 1:
                exit_price = float(ops[-1]['precio']) if ops[-1]['precio'] else None
            
            total_contracts = sum(float(op['contratos'] or 0) for op in ops if str(op['direccion']).upper() == str(first_op['direccion']).upper())
            
            # Determinar close_reason basado en PnL
            if total_pnl > 0:
                close_reason = 'PROFIT'
            elif total_pnl < 0:
                close_reason = 'LOSS'
            else:
                close_reason = 'BREAKEVEN'
            
            pnl_pct = ((exit_price - entry_price) / entry_price * 100) if entry_price and exit_price and side == 'buy' else \
                      ((entry_price - exit_price) / entry_price * 100) if entry_price and exit_price and side == 'sell' else None
            
            c.execute('''INSERT INTO trades 
                (trade_id, symbol, side, entry_price, exit_price,
                 leverage, amount_usdt, contracts, pnl_usdt, pnl_pct,
                 close_reason, opened_at, closed_at, commission, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'closed')''', (
                f'MIGRATED-{key}',
                symbol,
                side,
                entry_price,
                exit_price,
                int(str(leverage).replace('x', '')) if leverage else 10,
                10.0,
                total_contracts,
                total_pnl,
                pnl_pct,
                close_reason,
                fecha,
                fecha,  # Mismo timestamp ya que no sabemos la duración exacta
                total_comision,
            ))
        
        conn.commit()
        migrated = c.execute("SELECT COUNT(*) FROM trades").fetchone()[0]
        conn.close()
        
        print(f"✅ Migración completada: {migrated} trades importados a la DB")
        return True
        
    except Exception as e:
        print(f"❌ Error en migración: {e}")
        import traceback
        traceback.print_exc()
        return False


def _ensure_db_tables(db_path):
    """Crea las tablas de la Analytics DB si no existen."""
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS trades (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        trade_id        TEXT,
        symbol          TEXT NOT NULL,
        side            TEXT,
        entry_price     REAL,
        exit_price      REAL,
        sl_original     REAL,
        sl_final        REAL,
        tp1             REAL,
        tp2             REAL,
        tp3             REAL,
        hit_tp1         INTEGER DEFAULT 0,
        hit_tp2         INTEGER DEFAULT 0,
        hit_tp3         INTEGER DEFAULT 0,
        hit_sl          INTEGER DEFAULT 0,
        hit_be          INTEGER DEFAULT 0,
        close_reason    TEXT,
        confidence      REAL,
        trend_line      TEXT,
        timeframe       TEXT,
        exchange        TEXT,
        leverage        INTEGER DEFAULT 10,
        amount_usdt     REAL DEFAULT 10.0,
        contracts       REAL,
        pnl_usdt        REAL,
        pnl_pct         REAL,
        duration_min    REAL,
        opened_at       TEXT,
        closed_at       TEXT,
        commission      REAL,
        status          TEXT DEFAULT 'open'
    )''')
    c.execute('''CREATE TABLE IF NOT EXISTS trade_events (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        trade_id    TEXT,
        symbol      TEXT NOT NULL,
        action      TEXT NOT NULL,
        reason      TEXT,
        status      TEXT,
        price       REAL,
        sl          REAL,
        new_sl      REAL,
        close_pct   REAL,
        confidence  REAL,
        raw_payload TEXT,
        created_at  TEXT DEFAULT (datetime('now','localtime'))
    )''')
    conn.commit()
    conn.close()


# ════════════════════════════════════════════════════════════════
# MAIN — Ejecutar standalone
# ════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    print("═" * 60)
    print("  📊 Excel Manager — Registro de Trades v6.2")
    print("═" * 60)
    print()
    
    # 1. Asegurar que las tablas existen
    print("1️⃣  Verificando/creando tablas en Analytics DB...")
    _ensure_db_tables(ANALYTICS_DB)
    
    # 2. Migrar datos del Excel viejo si la DB está vacía
    print("2️⃣  Migrando datos existentes del Excel a la DB...")
    migrate_existing_excel_to_db()
    
    # 3. Exportar a Excel nuevo
    print("3️⃣  Exportando trades a Excel...")
    export_trades_to_excel()
    
    print()
    print("═" * 60)
    print("  ✅ Proceso completado")
    print("═" * 60)
